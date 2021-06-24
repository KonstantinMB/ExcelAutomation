# We use these import to create a new workbook OR load an existing one!
from openpyxl import Workbook, load_workbook

wb = load_workbook('investments.xlsx')
ws = wb.active  # this gives us the active worksheet(ws) from the workbook(wb)

# in case you want ws to not be the current active sheet, you can type the wb['name of sheet you want to work on']

# In order to change a certain value from the Excel file:
# ws['A2'] = "Testing"
# wb.save('investments.xlsx')

# In order to create a new sheet type:
# wb.create_sheet('Testing')

# If you want to access a specific sheet:
# wb['Sheet1'] where 'Sheet1' is the name of the sheet
