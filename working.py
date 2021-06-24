from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# now, lets create a new Workbook:
# wb = Workbook()
wb = load_workbook('me.xlsx')
ws = wb.active # gives as the default working sheet from the WorkBook
ws.title = "Data"

# In order to add data to the excel sheet, you can use a for-loop
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)  # uses the importing we gave it to set the row/col from 11 to A1 for example
        ws[char + str(row)] = char + str(row)

wb.save('me.xlsx')

