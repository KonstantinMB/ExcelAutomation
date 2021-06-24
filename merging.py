from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


wb = load_workbook('me.xlsx')
ws = wb.active

ws.merge_cells("A1:D1")  # using this command to merge the declared in brackets cells in one cell

# if you want to unmerge the cells:
# ws.unmerge_cells("A1:D1")  # BUT keep in mind that the data from the previously unmerged cells want come back

wb.save('me.xlsx')

